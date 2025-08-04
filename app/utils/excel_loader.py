import logging
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import update
from app.db.models import User  # Импортируйте вашу модель User

async def load_initial_data_from_excel(session, file_path: str):
    """
    Считывает Excel-файл и заполняет таблицу User,
    если она изначально пустая.

    Ожидается следующая структура столбцов:
    1. TG ID
    2. WP ID
    3. Username
    4. Имя
    5. Статус
    6. Дата регистрации
    7. Последняя активность
    8. Дата последнего посещения
    9. Просмотренные ключевые слова
    10. Последний просмотр
    11. Подписан на рассылки (по статусу)
    """
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        row_num = 2
        rows_added = 0

        while True:
            tg_id_cell = sheet.cell(row=row_num, column=1)
            if not tg_id_cell.value:
                break

            tg_id = str(tg_id_cell.value).strip()
            wp_id = str(sheet.cell(row=row_num, column=2).value or "").strip()
            username = str(sheet.cell(row=row_num, column=3).value or "").strip()
            name = str(sheet.cell(row=row_num, column=4).value or "").strip()
            status = str(sheet.cell(row=row_num, column=5).value or "").strip()
            reg_date = str(sheet.cell(row=row_num, column=6).value or "").strip()
            last_active = str(sheet.cell(row=row_num, column=7).value or "").strip()
            last_visit = str(sheet.cell(row=row_num, column=8).value or "").strip()
            keywords = str(sheet.cell(row=row_num, column=9).value or "").strip()
            last_keyword_view = str(sheet.cell(row=row_num, column=10).value or "").strip()
            subscribed = str(sheet.cell(row=row_num, column=11).value or "").strip()

            # Парсим дату регистрации, если она есть
            created_at = None
            if reg_date and reg_date != "—":
                try:
                    # Парсим дату в формате "dd.mm.yyyy hh:mm"
                    created_at = datetime.strptime(reg_date, "%d.%m.%Y %H:%M")
                except ValueError:
                    logging.warning(f"Не удалось распарсить дату регистрации: {reg_date}")
            
            # Парсим дату последней активности, если она есть
            last_interaction = None
            if last_active and last_active != "—":
                try:
                    last_interaction = datetime.strptime(last_active, "%d.%m.%Y %H:%M")
                except ValueError:
                    logging.warning(f"Не удалось распарсить дату последней активности: {last_active}")
            
            user_obj = User(
                tg_id=tg_id,
                wp_id=wp_id if wp_id and wp_id != "—" else None,
                username_in_tg=username.replace("@", "") if username and username != "—" else None,
                first_name=name if name and name != "—" else None,
                status=status if status and status != "—" else None,
                created_at=created_at,
                last_interaction=last_interaction
            )

            session.add(user_obj)
            rows_added += 1
            row_num += 1

        await session.commit()
        logging.info(f"Загрузка завершена: добавлено {rows_added} пользователь(ей).")

    except FileNotFoundError:
        logging.error(f"Файл '{file_path}' не найден.")
    except Exception as e:
        logging.error(f"Ошибка при загрузке Excel: {e}")
